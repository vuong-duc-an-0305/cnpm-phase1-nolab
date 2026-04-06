"""
Celery Tasks for async report generation
Phase 3: Background processing to avoid timeout on large reports
"""
from celery import shared_task
from datetime import datetime, date, timedelta
from io import BytesIO
import base64

from openpyxl import Workbook
from django.utils import timezone

from api.orders.services import OrderService


def _build_summary_sheet(wb, from_date_obj, to_date_obj):
    """Build summary sheet with revenue statistics"""
    ws = wb.active if wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active.title == "Sheet" else wb.create_sheet()
    ws.title = "Summary"
    stats = OrderService.get_revenue_statistics(from_date_obj, to_date_obj)
    summary = stats.get('summary', {})
    ws.append(["From", "To", "Total Orders", "Total Revenue", "Total Discount", "Average Order Value"])
    ws.append([
        from_date_obj.isoformat() if from_date_obj else "",
        to_date_obj.isoformat() if to_date_obj else "",
        summary.get('total_orders', 0),
        summary.get('total_revenue', 0),
        summary.get('total_discount', 0),
        summary.get('average_order_value', 0),
    ])


def _build_trend_sheet(wb, from_date_obj, to_date_obj, group_by: str):
    """Build trend sheet with revenue over time"""
    ws = wb.create_sheet(title="Trend")
    data = OrderService.get_revenue_trend(from_date_obj, to_date_obj, group_by)
    labels = data.get('labels', [])
    datasets = data.get('datasets') or []
    series = datasets[0] if datasets else {"data": []}
    values = series.get('data', [])
    ws.append(["Period", "Revenue"])
    for label, value in zip(labels, values):
        ws.append([label, value])


def _build_by_category_sheet(wb, from_date_obj, to_date_obj):
    """Build category sheet with revenue by category"""
    ws = wb.create_sheet(title="ByCategory")
    data = OrderService.get_revenue_by_category(from_date_obj, to_date_obj)
    labels = data.get('labels', [])
    values = (data.get('datasets') or [{}])[0].get('data', [])
    ws.append(["Category", "Revenue"])
    for label, value in zip(labels, values):
        ws.append([label, value])


def _build_best_selling_sheet(wb, from_date_obj, to_date_obj, limit: int):
    """Build best selling products sheet"""
    ws = wb.create_sheet(title="BestSelling")
    items = OrderService.get_best_selling_products(from_date_obj, to_date_obj, limit)
    ws.append(["ProductID", "ProductName", "TotalQuantity", "TotalRevenue", "OrderCount"])
    for item in items:
        ws.append([
            item.get('ProductID'),
            item.get('ProductID__ProductName'),
            item.get('total_quantity'),
            item.get('total_revenue'),
            item.get('order_count'),
        ])


@shared_task(bind=True, name='reports.generate_revenue_report')
def generate_revenue_report_task(self, from_date_str=None, to_date_str=None, period='today', report_type='full'):
    """
    Async task to generate revenue report Excel file
    
    Args:
        from_date_str: Start date (YYYY-MM-DD)
        to_date_str: End date (YYYY-MM-DD)
        period: Period shortcut (today|week|month|year)
        report_type: Type of report (revenue|trend|by_category|best_selling|full)
    
    Returns:
        dict: {
            'filename': str,
            'file_data': str (base64 encoded),
            'size': int (bytes)
        }
    """
    try:
        # Update task state
        self.update_state(state='PROGRESS', meta={'current': 10, 'total': 100, 'status': 'Parsing parameters...'})
        
        # Parse dates
        def _parse_date(s: str):
            return datetime.strptime(s, '%Y-%m-%d').date()
        
        from_date_obj = None
        to_date_obj = None
        
        if from_date_str:
            from_date_obj = _parse_date(from_date_str)
        if to_date_str:
            to_date_obj = _parse_date(to_date_str)
        
        # If no dates provided, use period
        if not from_date_obj and not to_date_obj:
            today = timezone.localdate()
            if period == 'today':
                from_date_obj = today
                to_date_obj = today
            elif period == 'week':
                start = today - timedelta(days=today.weekday())
                from_date_obj = start
                to_date_obj = today
            elif period == 'month':
                start = today.replace(day=1)
                from_date_obj = start
                to_date_obj = today
            elif period == 'year':
                start = date(today.year, 1, 1)
                from_date_obj = start
                to_date_obj = today
        
        # Create workbook
        self.update_state(state='PROGRESS', meta={'current': 30, 'total': 100, 'status': 'Creating workbook...'})
        wb = Workbook()
        
        # Build sheets based on report type
        if report_type == 'revenue':
            self.update_state(state='PROGRESS', meta={'current': 50, 'total': 100, 'status': 'Building revenue sheet...'})
            _build_summary_sheet(wb, from_date_obj, to_date_obj)
        elif report_type == 'trend':
            self.update_state(state='PROGRESS', meta={'current': 50, 'total': 100, 'status': 'Building trend sheet...'})
            _build_trend_sheet(wb, from_date_obj, to_date_obj, 'day')
        elif report_type == 'by_category':
            self.update_state(state='PROGRESS', meta={'current': 50, 'total': 100, 'status': 'Building category sheet...'})
            _build_by_category_sheet(wb, from_date_obj, to_date_obj)
        elif report_type == 'best_selling':
            self.update_state(state='PROGRESS', meta={'current': 50, 'total': 100, 'status': 'Building best selling sheet...'})
            _build_best_selling_sheet(wb, from_date_obj, to_date_obj, 10)
        elif report_type == 'full':
            self.update_state(state='PROGRESS', meta={'current': 40, 'total': 100, 'status': 'Building summary...'})
            _build_summary_sheet(wb, from_date_obj, to_date_obj)
            
            self.update_state(state='PROGRESS', meta={'current': 55, 'total': 100, 'status': 'Building trend...'})
            _build_trend_sheet(wb, from_date_obj, to_date_obj, 'day')
            
            self.update_state(state='PROGRESS', meta={'current': 70, 'total': 100, 'status': 'Building categories...'})
            _build_by_category_sheet(wb, from_date_obj, to_date_obj)
            
            self.update_state(state='PROGRESS', meta={'current': 85, 'total': 100, 'status': 'Building best selling...'})
            _build_best_selling_sheet(wb, from_date_obj, to_date_obj, 10)
        
        # Generate filename
        if from_date_str or to_date_str:
            start_str = from_date_obj.isoformat() if from_date_obj else ''
            end_str = to_date_obj.isoformat() if to_date_obj else ''
            suffix = f"{start_str}__{end_str}".strip('_')
        else:
            suffix = period
        filename = f"bao_cao_{suffix}.xlsx"
        
        # Save to BytesIO and encode as base64
        self.update_state(state='PROGRESS', meta={'current': 95, 'total': 100, 'status': 'Saving file...'})
        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        file_data = base64.b64encode(buffer.read()).decode('utf-8')
        file_size = buffer.tell()
        
        return {
            'filename': filename,
            'file_data': file_data,
            'size': file_size,
            'from_date': from_date_obj.isoformat() if from_date_obj else None,
            'to_date': to_date_obj.isoformat() if to_date_obj else None,
            'report_type': report_type
        }
        
    except Exception as e:
        # Update task state to FAILURE with error info
        self.update_state(
            state='FAILURE',
            meta={
                'exc_type': type(e).__name__,
                'exc_message': str(e),
                'current': 0,
                'total': 100,
                'status': f'Error: {str(e)}'
            }
        )
        raise


@shared_task(name='reports.cleanup_old_results')
def cleanup_old_results_task():
    """
    Periodic task to cleanup old Celery results from Redis
    Should be scheduled to run daily
    """
    from celery.result import AsyncResult
    from BE_coffee.celery import app
    
    # This is a placeholder - actual implementation would need
    # to iterate through result keys in Redis and delete old ones
    # For now, Celery's CELERY_RESULT_EXPIRES setting handles this
    
    return {'status': 'cleanup completed'}


@shared_task(name='reports.daily_summary_task')
def daily_summary_task():
    """
    Periodic task chạy vào cuối ngày để gửi tổng kết
    Chạy lúc 23:50 hàng ngày
    """
    from datetime import date
    from api.orders.services import OrderService
    from api.realtime.notifications import send_daily_summary
    from decimal import Decimal
    
    today = date.today()
    
    # Lấy thống kê ngày hôm nay
    stats = OrderService.get_revenue_statistics(
        from_date=today,
        to_date=today,
        include_all_status=True
    )
    
    if stats and 'summary' in stats:
        total_orders = stats['summary']['total_orders']
        total_revenue = stats['summary']['total_revenue']
        
        # Lấy top products
        top_products = []
        if 'daily_details' in stats and stats['daily_details']:
            daily_data = stats['daily_details'][0]  # Chỉ có 1 ngày
            if 'top_products' in daily_data:
                top_products = daily_data['top_products'][:5]  # Top 5
        
        # Gửi tổng kết
        send_daily_summary(
            date=today,
            total_orders=total_orders,
            total_revenue=total_revenue,
            top_products=top_products
        )
        
        return {
            'status': 'success',
            'date': today.isoformat(),
            'total_orders': total_orders,
            'total_revenue': float(total_revenue)
        }
    
    return {'status': 'no_data'}


@shared_task(name='reports.check_low_stock_task')
def check_low_stock_task():
    """
    Periodic task kiểm tra hàng tồn kho thấp
    Chạy mỗi 6 giờ
    """
    from api.ingredients.models import NguyenLieu
    from api.realtime.notifications import send_low_stock_alert
    
    low_stock_count = 0
    
    # Lấy tất cả nguyên liệu có MinQuantity
    ingredients = NguyenLieu.objects.filter(
        MinQuantity__gt=0,
        IsActive=True
    ).only('IngredientID', 'IngredientName', 'Quantity', 'MinQuantity')
    
    for ingredient in ingredients:
        # Alert nếu tồn kho < 150% min_quantity
        if ingredient.Quantity < ingredient.MinQuantity * 1.5:
            send_low_stock_alert(
                ingredient_name=ingredient.IngredientName,
                current_quantity=ingredient.Quantity,
                min_quantity=ingredient.MinQuantity,
                ingredient_id=ingredient.IngredientID
            )
            low_stock_count += 1
    
    return {
        'status': 'success',
        'low_stock_items': low_stock_count,
        'total_checked': ingredients.count()
    }
