"""
Views xuất báo cáo XLSX (streaming)
"""
from datetime import datetime, date, timedelta
from io import BytesIO

from django.utils.encoding import iri_to_uri
from django.http import StreamingHttpResponse
from django.utils import timezone

from rest_framework.decorators import api_view, permission_classes
from rest_framework import status
from rest_framework.response import Response

from openpyxl import Workbook

from api.orders.services import OrderService
from api.core.permissions import IsAdminRole


def _build_summary_sheet(wb, from_date_obj, to_date_obj):
    ws = wb.active if wb.active.max_row == 1 and wb.active.max_column == 1 and wb.active.title == "Sheet" else wb.create_sheet()
    ws.title = "Summary"
    stats = OrderService.get_revenue_statistics(from_date_obj, to_date_obj)
    summary = stats.get('summary', {})
    ws.append(["From", "To", "Total Orders", "Total Revenue", "Total Discount", "Average Order Value"])
    ws.append([
        from_date_obj.isoformat() if from_date_obj else "",
        to_date_obj.isoformat() if to_date_obj else "",
        summary.get('total_orders', 0),
        summary.get('total_revenue', 0),
        summary.get('total_discount', 0),
        summary.get('average_order_value', 0),
    ])


def _build_trend_sheet(wb, from_date_obj, to_date_obj, group_by: str):
    ws = wb.create_sheet(title="Trend")
    data = OrderService.get_revenue_trend(from_date_obj, to_date_obj, group_by)
    labels = data.get('labels', [])
    datasets = data.get('datasets') or []
    series = datasets[0] if datasets else {"data": []}
    values = series.get('data', [])
    ws.append(["Period", "Revenue"])
    for label, value in zip(labels, values):
        ws.append([label, value])


def _build_by_category_sheet(wb, from_date_obj, to_date_obj):
    ws = wb.create_sheet(title="ByCategory")
    data = OrderService.get_revenue_by_category(from_date_obj, to_date_obj)
    labels = data.get('labels', [])
    values = (data.get('datasets') or [{}])[0].get('data', [])
    ws.append(["Category", "Revenue"])
    for label, value in zip(labels, values):
        ws.append([label, value])


def _build_best_selling_sheet(wb, from_date_obj, to_date_obj, limit: int):
    ws = wb.create_sheet(title="BestSelling")
    items = OrderService.get_best_selling_products(from_date_obj, to_date_obj, limit)
    ws.append(["ProductID", "ProductName", "TotalQuantity", "TotalRevenue", "OrderCount"])
    for item in items:
        ws.append([
            item.get('ProductID'),
            item.get('ProductID__ProductName'),
            item.get('total_quantity'),
            item.get('total_revenue'),
            item.get('order_count'),
        ])


def _workbook_to_chunks(wb, chunk_size: int = 4096):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    while True:
        data = buffer.read(chunk_size)
        if not data:
            break
        yield data


@api_view(['GET'])
@permission_classes([IsAdminRole])
def export_report_xlsx(request):
    """
    GET /api/reports/export.xlsx?from_date&to_date&type=revenue
    Trả về streaming XLSX: Content-Type + Content-Disposition
    Tên file: bao_cao_{ngày xuất báo cáo}.xlsx
    """
    report_type = (request.query_params.get('type') or 'revenue').lower()
    group_by = (request.query_params.get('group_by') or 'day').lower()
    period = (request.query_params.get('period') or '').lower()
    limit_param = request.query_params.get('limit')
    try:
        best_limit = int(limit_param) if limit_param is not None else 10
    except ValueError:
        best_limit = 10

    from_date = request.query_params.get('from_date')
    to_date = request.query_params.get('to_date')

    # Chuẩn hóa phạm vi ngày: ưu tiên from_date/to_date, nếu không có thì dùng period
    def _parse_date(s: str):
        return datetime.strptime(s, '%Y-%m-%d').date()

    from_date_obj = None
    to_date_obj = None
    if from_date or to_date:
        if from_date:
            try:
                from_date_obj = _parse_date(from_date)
            except ValueError:
                return Response({'error': 'from_date phải có định dạng YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
        if to_date:
            try:
                to_date_obj = _parse_date(to_date)
            except ValueError:
                return Response({'error': 'to_date phải có định dạng YYYY-MM-DD'}, status=status.HTTP_400_BAD_REQUEST)
    else:
        # Map period -> from_date/to_date theo timezone hiện tại
        today = timezone.localdate()
        if period in ('today', ''):
            from_date_obj = today
            to_date_obj = today
        elif period == 'week':
            start = today - timedelta(days=today.weekday())
            from_date_obj = start
            to_date_obj = today
        elif period == 'month':
            start = today.replace(day=1)
            from_date_obj = start
            to_date_obj = today
        elif period == 'year':
            start = date(today.year, 1, 1)
            from_date_obj = start
            to_date_obj = today
        else:
            # period không hợp lệ
            return Response({'error': 'period không hợp lệ. Hỗ trợ: today|week|month|year hoặc dùng from_date/to_date'}, status=status.HTTP_400_BAD_REQUEST)

    wb = Workbook()

    # Xây sheet theo type
    if report_type == 'revenue':
        _build_summary_sheet(wb, from_date_obj, to_date_obj)
    elif report_type == 'trend':
        _build_trend_sheet(wb, from_date_obj, to_date_obj, group_by)
    elif report_type == 'by_category':
        _build_by_category_sheet(wb, from_date_obj, to_date_obj)
    elif report_type == 'best_selling':
        _build_best_selling_sheet(wb, from_date_obj, to_date_obj, best_limit)
    elif report_type == 'full':
        _build_summary_sheet(wb, from_date_obj, to_date_obj)
        _build_trend_sheet(wb, from_date_obj, to_date_obj, group_by)
        _build_by_category_sheet(wb, from_date_obj, to_date_obj)
        _build_best_selling_sheet(wb, from_date_obj, to_date_obj, best_limit)
    else:
        ws = wb.active
        ws.title = "Report"
        ws.append(["Message"]) 
        ws.append([f"Unsupported report type: {report_type}"])

    # Đặt tên file: bao_cao_{period|from_to}.xlsx
    if request.query_params.get('from_date') or request.query_params.get('to_date'):
        start_str = from_date_obj.isoformat() if from_date_obj else ''
        end_str = to_date_obj.isoformat() if to_date_obj else ''
        suffix = f"{start_str}__{end_str}".strip('_')
    else:
        suffix = (period or 'today')

    filename = f"bao_cao_{suffix}.xlsx"

    response = StreamingHttpResponse(
        _workbook_to_chunks(wb),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{iri_to_uri(filename)}"'
    return response


@api_view(['POST'])
@permission_classes([IsAdminRole])
def export_report_async(request):
    """
    POST /api/reports/export_async/
    Start async report generation task
    
    Body: {
        "from_date": "2024-01-01",  # optional
        "to_date": "2024-12-31",    # optional
        "period": "today|week|month|year",  # optional, default: today
        "report_type": "revenue|trend|by_category|best_selling|full"  # optional, default: full
    }
    
    Returns: {
        "task_id": "uuid",
        "status": "PENDING",
        "message": "Task started"
    }
    """
    from api.reports.tasks import generate_revenue_report_task
    
    from_date = request.data.get('from_date')
    to_date = request.data.get('to_date')
    period = request.data.get('period', 'today')
    report_type = request.data.get('report_type', 'full')
    
    # Validate report_type
    valid_types = ['revenue', 'trend', 'by_category', 'best_selling', 'full']
    if report_type not in valid_types:
        return Response(
            {'error': f'report_type phải là một trong: {", ".join(valid_types)}'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    # Start async task
    task = generate_revenue_report_task.delay(
        from_date_str=from_date,
        to_date_str=to_date,
        period=period,
        report_type=report_type
    )
    
    return Response({
        'task_id': task.id,
        'status': task.status,
        'message': 'Async report generation started'
    }, status=status.HTTP_202_ACCEPTED)


@api_view(['GET'])
@permission_classes([IsAdminRole])
def check_task_status(request, task_id):
    """
    GET /api/reports/task_status/<task_id>/
    Check status of async task
    
    Returns: {
        "task_id": "uuid",
        "status": "PENDING|PROGRESS|SUCCESS|FAILURE",
        "result": {...} or null,
        "meta": {...}
    }
    """
    from celery.result import AsyncResult
    
    task = AsyncResult(task_id)
    
    response_data = {
        'task_id': task_id,
        'status': task.status,
    }
    
    if task.status == 'PENDING':
        response_data['message'] = 'Task is waiting to be executed'
    elif task.status == 'PROGRESS':
        response_data['meta'] = task.info
        response_data['message'] = task.info.get('status', 'Processing...')
    elif task.status == 'SUCCESS':
        response_data['result'] = task.result
        response_data['message'] = 'Task completed successfully'
    elif task.status == 'FAILURE':
        response_data['error'] = str(task.info)
        response_data['message'] = 'Task failed'
    
    return Response(response_data)


@api_view(['GET'])
@permission_classes([IsAdminRole])
def download_task_result(request, task_id):
    """
    GET /api/reports/download/<task_id>/
    Download completed report file
    
    Returns: Excel file or error
    """
    from celery.result import AsyncResult
    import base64
    from django.http import HttpResponse
    
    task = AsyncResult(task_id)
    
    if task.status != 'SUCCESS':
        return Response(
            {'error': f'Task not completed. Current status: {task.status}'},
            status=status.HTTP_400_BAD_REQUEST
        )
    
    result = task.result
    filename = result.get('filename', 'report.xlsx')
    file_data = result.get('file_data')
    
    if not file_data:
        return Response(
            {'error': 'No file data available'},
            status=status.HTTP_404_NOT_FOUND
        )
    
    # Decode base64 file data
    file_bytes = base64.b64decode(file_data)
    
    response = HttpResponse(
        file_bytes,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{iri_to_uri(filename)}"'
    response['Content-Length'] = len(file_bytes)
    
    return response


